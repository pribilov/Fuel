import streamlit as st
import math, io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

st.set_page_config(page_title="Fuel Order", page_icon="✈️", layout="wide")

st.markdown("""
<style>
.block-container { padding-top: 1rem; }
.sec-header {
    background-color: #003087; color: white; font-weight: bold;
    font-size: 13px; text-align: center; padding: 5px 0;
    border: 1px solid #003087; margin-bottom: 0;
}
.sec-header-light {
    background-color: #e8eef7; color: #003087; font-weight: bold;
    font-size: 12px; text-align: center; padding: 4px 0;
    border: 1px solid #bbc8dc;
}
.flabel {
    font-size: 10px; font-weight: bold; color: #555;
    text-transform: uppercase; margin-bottom: 2px;
}
.bigval {
    font-size: 22px; font-weight: 900; color: #003087;
    text-align: right; border-bottom: 2px solid #003087; padding-bottom: 2px;
}
.bigval-auto {
    font-size: 22px; font-weight: 900; color: #888;
    text-align: right; border-bottom: 2px dashed #aaa; padding-bottom: 2px;
}
.diff-pos  { font-size: 26px; font-weight: 900; color: #1a7a1a; text-align:center; }
.diff-neg  { font-size: 26px; font-weight: 900; color: #c0392b; text-align:center; }
.diff-zero { font-size: 26px; font-weight: 900; color: #003087; text-align:center; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════
h1, h2 = st.columns([1, 2])
with h1:
    st.markdown("### ✈️ LATAM AIRLINES")
with h2:
    st.markdown(
        '<div style="font-size:32px;font-weight:900;color:#003087;'
        'border:3px solid #003087;text-align:center;padding:6px 0;'
        'letter-spacing:4px;">FUEL ORDER</div>',
        unsafe_allow_html=True
    )
st.markdown("<hr style='border:2px solid #003087;margin:8px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# STATION / DATE / FLIGHT / FROM-TO / A/C REG
# ═══════════════════════════════════════════════
c1, c2, c3, c4, c5 = st.columns([1, 1.3, 1, 1.8, 1.2])
with c1:
    st.markdown('<p class="flabel">Station</p>', unsafe_allow_html=True)
    station = st.text_input("_station", label_visibility="collapsed")
with c2:
    st.markdown('<p class="flabel">Date</p>', unsafe_allow_html=True)
    fl_date = st.date_input("_date", value=date.today(), label_visibility="collapsed")
with c3:
    st.markdown('<p class="flabel">Flight</p>', unsafe_allow_html=True)
    flight = st.text_input("_flight", label_visibility="collapsed")
with c4:
    st.markdown('<p class="flabel">From / To</p>', unsafe_allow_html=True)
    from_to = st.text_input("_fromto", placeholder="LIM  →  LAX", label_visibility="collapsed")
with c5:
    st.markdown('<p class="flabel">A/C Reg</p>', unsafe_allow_html=True)
    ac_reg = st.text_input("_acreg", label_visibility="collapsed")

st.markdown('<p class="flabel" style="margin-top:6px">Prepared by — Flight Operations Off/SM Name</p>',
            unsafe_allow_html=True)
prepared_by = st.text_input("_prepby", label_visibility="collapsed")

st.markdown("<hr style='border:1px solid #003087;margin:8px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# DENSITIES
# ═══════════════════════════════════════════════
dc1, dc2, dc3, dc4, dc5 = st.columns([1.4, 1.1, 1.2, 1.4, 1.4])
with dc1:
    st.markdown('<p class="flabel">Densities (kg/lt)</p>', unsafe_allow_html=True)
    density = st.number_input("_dens", min_value=0.001, max_value=1.0,
                              value=0.800, step=0.001, format="%.3f",
                              label_visibility="collapsed")
with dc2:
    st.markdown('<p class="flabel">γ standard (ref)</p>', unsafe_allow_html=True)
    st.markdown('<div class="bigval-auto">0.600</div>', unsafe_allow_html=True)
with dc3:
    gamma = density * 3.786
    st.markdown('<p class="flabel">γ Suppl (auto)</p>', unsafe_allow_html=True)
    st.markdown(f'<div class="bigval">{gamma:.3f}</div>', unsafe_allow_html=True)
with dc4:
    st.markdown('<p class="flabel">Time</p>', unsafe_allow_html=True)
    dens_time = st.text_input("_denstime", label_visibility="collapsed")
with dc5:
    st.markdown('<p class="flabel">Signature</p>', unsafe_allow_html=True)
    dens_sig = st.text_input("_denssig", label_visibility="collapsed")

st.markdown("<hr style='border:2px solid #003087;margin:8px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# MAIN BODY — FUEL KG  |  CAPTAIN OR STATION
# ═══════════════════════════════════════════════
left, right = st.columns([3, 2], gap="medium")

with right:
    st.markdown('<div class="sec-header">CAPTAIN OR STATION — KG</div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown('<p class="flabel">Refuel / Block Fuel (captain request)</p>',
                    unsafe_allow_html=True)
        captain = st.number_input("_captain", min_value=0, value=0, step=10,
                                  label_visibility="collapsed",
                                  help="El capitán puede pedir más que el block fuel del coordinador.")
        rc1, rc2 = st.columns(2)
        with rc1:
            st.markdown('<p class="flabel">Time</p>', unsafe_allow_html=True)
            capt_time = st.text_input("_capttime", label_visibility="collapsed")
        with rc2:
            st.markdown('<p class="flabel">Name 1</p>', unsafe_allow_html=True)
            capt_name = st.text_input("_captname", label_visibility="collapsed")
        st.markdown('<p class="flabel">Signature</p>', unsafe_allow_html=True)
        capt_sig = st.text_input("_captsig", label_visibility="collapsed")

with left:
    st.markdown('<div class="sec-header">FUEL KG</div>', unsafe_allow_html=True)
    with st.container(border=True):
        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">Take Off Fuel</p>', unsafe_allow_html=True)
        with r2:
            takeoff_fuel = st.number_input("_tof", min_value=0, value=0, step=10,
                                           label_visibility="collapsed")

        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">Taxi Fuel  ＋</p>', unsafe_allow_html=True)
        with r2:
            taxi_fuel = st.number_input("_taxi", min_value=0, value=0, step=10,
                                        label_visibility="collapsed")

        block_fuel = takeoff_fuel + taxi_fuel
        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">Block Fuel <span style="color:#aaa">(auto)</span></p>',
                        unsafe_allow_html=True)
        with r2:
            st.markdown(f'<div class="bigval-auto">{block_fuel:,}</div>', unsafe_allow_html=True)

        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">Remain  －</p>', unsafe_allow_html=True)
        with r2:
            remain = st.number_input("_remain", min_value=0, value=0, step=10,
                                     label_visibility="collapsed")

    # Computed values
    to_be_kg  = captain - remain
    to_be_lts = to_be_kg / density if density > 0 else 0
    gal_exact = to_be_kg / gamma   if gamma   > 0 else 0
    gal_up    = math.ceil(gal_exact / 100) * 100

    with st.container(border=True):
        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">To Be Refueled (kg)</p>', unsafe_allow_html=True)
        with r2:
            st.markdown(f'<div class="bigval">{to_be_kg:,}</div>', unsafe_allow_html=True)

        st.markdown('<p style="font-size:11px;color:#999;text-align:center">↓  Kgs ÷ γ = LTS</p>',
                    unsafe_allow_html=True)

        r1, r2 = st.columns([2, 1])
        with r1:
            st.markdown('<p class="flabel">To Be Refueled (LTS)</p>', unsafe_allow_html=True)
        with r2:
            st.markdown(f'<div class="bigval">{to_be_lts:,.1f}</div>', unsafe_allow_html=True)

st.markdown("<hr style='border:2px solid #003087;margin:8px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# LTS / USG — TRUCK READING
# ═══════════════════════════════════════════════
st.markdown('<div class="sec-header-light">LTS / USG</div>', unsafe_allow_html=True)

s1, s2, s3, s4 = st.columns([2, 1.2, 1.2, 1.8])
with s1:
    st.markdown('<p class="flabel">Fuel Slips</p>', unsafe_allow_html=True)
    fuel_slips = st.text_input("_slips", placeholder="optional notes", label_visibility="collapsed")
with s2:
    st.markdown('<p class="flabel">To Request (gal ↑100)</p>', unsafe_allow_html=True)
    st.markdown(f'<div class="bigval">{gal_up:,}</div>', unsafe_allow_html=True)
    st.caption(f"exact: {gal_exact:,.1f} gal")
with s3:
    st.markdown('<p class="flabel">Truck Screen (USG)</p>', unsafe_allow_html=True)
    gal_truck = st.number_input("_galtruck", min_value=0, value=0, step=1,
                                label_visibility="collapsed")
with s4:
    fueled_lts = gal_truck * 3.786
    st.markdown('<p class="flabel">Total  (USG × 3.786 = LTS)</p>', unsafe_allow_html=True)
    st.markdown(f'<div class="bigval">{fueled_lts:,.1f} LTS</div>', unsafe_allow_html=True)

st.markdown("<hr style='border:1px solid #ccc;margin:8px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# TO BE REFUELED vs TOTAL FUELING + DIFFERENCE
# ═══════════════════════════════════════════════
t1, t2, gap, t3 = st.columns([1.5, 1.5, 0.3, 2])
with t1:
    st.markdown('<p class="flabel">To Be Refueled (LTS)</p>', unsafe_allow_html=True)
    st.markdown(f'<div class="bigval">{to_be_lts:,.1f} LTS</div>', unsafe_allow_html=True)
with t2:
    st.markdown('<p class="flabel">Total Fueling (LTS)</p>', unsafe_allow_html=True)
    st.markdown(f'<div class="bigval">{fueled_lts:,.1f} LTS</div>', unsafe_allow_html=True)
with t3:
    diff = to_be_lts - fueled_lts
    sign = "+" if diff > 0 else ""
    css  = "diff-zero" if diff == 0 else ("diff-pos" if diff >= 0 else "diff-neg")
    st.markdown('<div class="sec-header">DIFFERENCE OF FUELED</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="{css}">{sign}{diff:,.1f} LTS</div>', unsafe_allow_html=True)
    if diff < 0:
        st.warning(f"⚠️ {abs(diff):,.1f} L de más")
    elif diff > 0:
        st.info(f"ℹ️ {diff:,.1f} L pendientes")
    else:
        st.success("✅ Carga exacta")

st.markdown("<hr style='border:2px solid #003087;margin:10px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# SIGNATURES
# ═══════════════════════════════════════════════
sg1, sg2 = st.columns(2)
with sg1:
    st.markdown('<div class="sec-header-light">Fueled by — Station Mechanic</div>',
                unsafe_allow_html=True)
    sc1, sc2 = st.columns(2)
    with sc1:
        st.markdown('<p class="flabel">Name</p>', unsafe_allow_html=True)
        mech_name = st.text_input("_mechname", label_visibility="collapsed")
    with sc2:
        st.markdown('<p class="flabel">Signature</p>', unsafe_allow_html=True)
        mech_sig = st.text_input("_mechsig", label_visibility="collapsed")
with sg2:
    st.markdown('<div class="sec-header-light">Checked by Captain</div>',
                unsafe_allow_html=True)
    sc1, sc2 = st.columns(2)
    with sc1:
        st.markdown('<p class="flabel">Name</p>', unsafe_allow_html=True)
        check_name = st.text_input("_checkname", label_visibility="collapsed")
    with sc2:
        st.markdown('<p class="flabel">Signature</p>', unsafe_allow_html=True)
        check_sig = st.text_input("_checksig", label_visibility="collapsed")

st.markdown("<hr style='border:2px solid #003087;margin:12px 0'>", unsafe_allow_html=True)

# ═══════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════
def build_excel(data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fuel Order"

    thin  = Side(style="thin",   color="BFBFBF")
    thick = Side(style="medium", color="003087")
    brd   = Border(left=thin,  right=thin,  top=thin,  bottom=thin)
    brd_h = Border(left=thick, right=thick, top=thick, bottom=thick)

    hdr_f = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    sec_f = Font(name="Arial", bold=True, color="003087", size=10)
    lbl_f = Font(name="Arial", size=9, italic=True)
    val_f = Font(name="Arial", bold=True, size=11)
    frm_f = Font(name="Arial", color="000000", size=11)
    hdr_fill = PatternFill("solid", fgColor="003087")
    sec_fill = PatternFill("solid", fgColor="BDD7EE")
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    ctr  = Alignment(horizontal="center", vertical="center")
    lft  = Alignment(horizontal="left",   vertical="center")
    rgt  = Alignment(horizontal="right",  vertical="center")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    def hdr(r, txt):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value, c.font, c.fill, c.alignment = txt, hdr_f, hdr_fill, ctr
        ws.row_dimensions[r].height = 22

    def sec(r, txt):
        ws.merge_cells(f"A{r}:D{r}")
        c = ws[f"A{r}"]
        c.value, c.font, c.fill, c.alignment = txt, sec_f, sec_fill, lft
        ws.row_dimensions[r].height = 18

    def put(r, a, b, c="", d="", fb=False, fd=False, fmtb=None, fmtd=None, shade=False):
        fill = alt_fill if shade else None
        for col in "ABCD":
            ws[f"{col}{r}"].border = brd
            if fill: ws[f"{col}{r}"].fill = fill
        ws[f"A{r}"].value, ws[f"A{r}"].font, ws[f"A{r}"].alignment = a, lbl_f, lft
        bv = ws[f"B{r}"]
        bv.value, bv.font, bv.alignment = b, (frm_f if fb else val_f), rgt
        if fmtb: bv.number_format = fmtb
        ws[f"C{r}"].value, ws[f"C{r}"].font, ws[f"C{r}"].alignment = c, lbl_f, lft
        dv = ws[f"D{r}"]
        dv.value, dv.font, dv.alignment = d, (frm_f if fd else val_f), rgt
        if fmtd: dv.number_format = fmtd
        ws.row_dimensions[r].height = 16

    d = data
    hdr(1, "✈  LATAM AIRLINES  —  FUEL ORDER")
    sec(3, "FLIGHT INFORMATION")
    put(4,  "Station",          d["station"],    "Date",       str(d["fl_date"]),  fmtb="@", fmtd="@")
    put(5,  "Flight",           d["flight"],     "From / To",  d["from_to"],       fmtb="@", fmtd="@", shade=True)
    put(6,  "A/C Reg",          d["ac_reg"],     "Prepared by",d["prepared_by"],   fmtb="@", fmtd="@")

    sec(8, "DENSITIES")
    put(9,  "Density (kg/lt)",  d["density"],    "γ standard", 0.600,  fmtb="0.000", fmtd="0.000", shade=True)
    put(10, "γ Suppl (auto)",  "=B9*3.786",      "Time",       d["dens_time"], fb=True, fmtb="0.000", fmtd="@")

    sec(12, "FUEL KG")
    put(13, "Take Off Fuel (kg)",   d["takeoff_fuel"], fmtb="#,##0")
    put(14, "Taxi Fuel + (kg)",     d["taxi_fuel"],    fmtb="#,##0", shade=True)
    put(15, "Block Fuel (kg)",      "=B13+B14",        fb=True, fmtb="#,##0")
    put(16, "Remain − (kg)",        d["remain"],       fmtb="#,##0", shade=True)

    sec(18, "CAPTAIN OR STATION")
    put(19, "Block Fuel — Captain", d["captain"],   "Time",      d["capt_time"], fmtb="#,##0", fmtd="@")
    put(20, "Name",                 d["capt_name"], "Signature", d["capt_sig"],  fmtb="@",     fmtd="@", shade=True)

    sec(22, "TO BE REFUELED")
    put(23, "To Be Refueled (kg)",   "=B19-B16",              fb=True, fmtb="#,##0", shade=True)
    put(24, "To Be Refueled (LTS)",  "=B23/B9",               fb=True, fmtb="#,##0.0")
    put(25, "Gallons to Request (↑100)", "=CEILING(B23/(B9*3.786),100)", fb=True, fmtb="#,##0", shade=True)

    sec(27, "TRUCK READING  —  LTS/USG")
    put(28, "Fuel Slips",           d["fuel_slips"], "Truck Screen (USG)", d["gal_truck"], fmtb="@", fmtd="#,##0", shade=True)
    put(29, "Total Fueling (LTS)",  "=D28*3.786",    "Diff of Fueled (LTS)", "=B24-B29", fb=True, fd=True, fmtb="#,##0.0", fmtd="#,##0.0")

    sec(31, "SIGNATURES")
    put(32, "Fueled by (Name)",  d["mech_name"],  "Fueled by (Sig)",  d["mech_sig"],  fmtb="@", fmtd="@", shade=True)
    put(33, "Checked by (Name)", d["check_name"], "Checked by (Sig)", d["check_sig"], fmtb="@", fmtd="@")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


st.subheader("📥 Exportar")
fname = f"FuelOrder_{flight or 'flight'}_{fl_date}.xlsx".replace(" ", "_")
st.download_button(
    label="⬇️  Descargar Excel",
    data=build_excel({
        "station": station, "fl_date": fl_date, "flight": flight,
        "from_to": from_to, "ac_reg": ac_reg, "prepared_by": prepared_by,
        "density": density, "dens_time": dens_time, "dens_sig": dens_sig,
        "takeoff_fuel": takeoff_fuel, "taxi_fuel": taxi_fuel,
        "captain": captain, "remain": remain,
        "gal_truck": gal_truck, "fuel_slips": fuel_slips,
        "capt_time": capt_time, "capt_name": capt_name, "capt_sig": capt_sig,
        "mech_name": mech_name, "mech_sig": mech_sig,
        "check_name": check_name, "check_sig": check_sig,
    }),
    file_name=fname,
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
